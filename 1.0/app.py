import requests
import random
import time
from openpyxl import load_workbook
from PyPDF2 import PdfReader

wb = load_workbook(filename = 'source.xlsx')
ws = wb.active

titles = []
pdf_filenames = []

###   from range 100 to 900 
###   create rows in sheet
###   check for pdf and save 

for count, cell in enumerate(ws['A']):
    url = cell.value
    response = requests.get(url)

    if response.status_code == 200:
        pdf_name = 'pdf/page_{}.pdf'.format(count + 1)
        pdf_filenames.append(pdf_name)

        with open(pdf_name, 'wb') as f:
            f.write(response.content)
            
        reader = PdfReader(pdf_name)
        meta = reader.metadata
        titles.append(meta.title)

    time.sleep(random.randint(2, 5))

for count, cell in enumerate(ws['B']):
    cell.value = titles[count]

for count, cell in enumerate(ws['C']):
    cell.value = pdf_filenames[count]

wb.save('output.xlsx')