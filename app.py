import requests
import random
import time
from openpyxl import Workbook
from PyPDF2 import PdfReader

base_url = 'https://www.diabetesresearch.org/document.doc?id='

wb = Workbook()
ws = wb.active

doc_ids = [*range(200, 300, 1)]

for count, id_num in enumerate(doc_ids):
    url = base_url + str(id_num)
    response = requests.get(url)

    if response.status_code == 200:
        # Save URL to spreadsheet
        ws['A{}'.format(count + 1)] = url

        pdf_name = 'pdf/{}.pdf'.format(id_num)

        # Save filename to spreadsheet
        ws['B{}'.format(count + 1)] = pdf_name

        # Download PDF document
        with open(pdf_name, 'wb') as f:
            f.write(response.content)
        
        # Get PDF metadata
        try:
            reader = PdfReader(pdf_name)
            meta = reader.metadata
        
            # Save PDF title to spreadsheet
            ws['C{}'.format(count + 1)] = meta.title
        except:
            # Record error
            ws['C{}'.format(count + 1)] = 'ERROR: PDF could not be opened'
            pass

    time.sleep(random.randint(1, 3))

# Save and output spreadsheet
wb.save('output.xlsx')